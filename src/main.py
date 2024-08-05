import csv
import chardet
from openpyxl import Workbook
import os

#list of files to analyse
path = '.'
try:
    files = os.listdir(path)
except OSError as e:
    print(f"Error accessing directory: {e}")
    files = []

csv_files = [name for name in files if name.endswith('.csv')]

if not csv_files:
    print("No CSV files found in the directory.")
else:
    print(csv_files)

#create excel file
wb = Workbook()
XLSXfilename = "data.xlsx"

#compute each file
for CSVfilename in csv_files:

    #infos files
    
    XLSXsheetname = CSVfilename 

    col_start = 1
    row_start = 1

    
    wb.create_sheet(XLSXsheetname)
    ws1 = wb[XLSXsheetname]

    # Detect encoding
    with open(CSVfilename, 'rb') as f:
        result = chardet.detect(f.read())
        currEncoding = result['encoding']

    #open file
    with open(CSVfilename, encoding=currEncoding) as csvfile:

        spamreader = csv.reader(csvfile, delimiter=';')
        
        for row in spamreader:
            #print(' '.join(row))
            
            liste = list(row)
            for var in liste:

                #write data to excel file
                ws1.cell(row_start, col_start).value = var
                col_start+=1
            col_start = 1
            row_start+=1
            
                

wb.save(XLSXfilename)
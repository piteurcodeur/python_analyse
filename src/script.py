
from dataclasses import dataclass
import glob
import os
from pathlib import Path
import numpy as np
from openpyxl import load_workbook
import openpyxl
import pandas as pd
from colorama import Fore, Style
import xlrd
import xlwt

EXIT_SUCCESS = 0
EXIT_FAILED = 1

"""

1) lire le fichier 1
2) chercher le fichier 2 et 3
3) si fichiers 2 et 3 existent : remplir le buffer avec le tableau du fichier 1
4) copier les fichiers 2 et 3 vers un buffer -> ecrire dans un fichier xlsx + ecrire le tableau buffer -> convertir en xls
5) sinon ecire le nom dans missing file txt


"""



@dataclass
class Filename :
    filename: str
    path: str
    number: int
    serialNumber: int



class Document :

    def __init__(self) -> None:
        current_script_dir = Path(__file__).parent
        self.Strang1Path = current_script_dir.parent / "Strang1/xls/"
        self.Strang2Path = current_script_dir.parent / "Strang2/xls/"
        self.Strang3Path = current_script_dir.parent / "Strang3/xls/"
        #self.Buffer = [[0] * 5 for _ in range(10)]


    def create_textfile(self) -> int:
        try:
            # Obtenez le répertoire du script actuel
            current_script_dir = Path(__file__).parent
            
            # Construisez le chemin du dossier cible
            output_dir = current_script_dir.parent / "output"
            
            # Assurez-vous que le dossier cible existe
            output_dir.mkdir(parents=True, exist_ok=True)
            
            # Chemin complet du fichier à créer
            self.missing_ref_textfile = output_dir / "missing.txt"
            
            # Créez le fichier (ou ouvrez-le s'il existe déjà) avec la méthode 'with'
            with open(self.missing_ref_textfile, "w+", encoding="UTF-8") as myFile:
                pass
            
            return EXIT_SUCCESS
        except Exception as e:
            print(f"Error: creating text file - {e}")
            return EXIT_FAILED


    def write_missing_ref(self, filename, numero) -> int:
        try:
            with open(self.missing_ref_textfile, "a") as myFile:
                myFile.write("Ref missing : " + filename.split('\\')[-1] + f" in file {numero}\n")

            return EXIT_SUCCESS
        except:
            print("Error : writing in missing ref txt file\n")
            return EXIT_FAILED
            

    def parse_name(self, name) -> int:
        try:
            lst = name.split("_")
            Filename.number = lst[-2]
            Filename.filename = name
            Filename.serialNumber = (lst[-1])[:-4][-4:]
        except:
            print("Error parsing name")
            return EXIT_FAILED
        return EXIT_SUCCESS


    def display_info(self) -> None:
        print("\n________Infos________\n")
        print("|__Filename: ", Filename.filename)
        print("|__Number: ", Filename.number)
        print("|__SerialNumber: ", Filename.serialNumber)
        print("______________________\n")


    def search_file(self, chemin) -> list:
        """
        Cherche un fichier .xls avec les patterns du fichier1 dans le <chemin> précisé
        Retourne le chemin local du fichier
        """
        pattern = Filename.serialNumber
        number = Filename.number
        #fichiers = [f for f in glob.glob(chemin + '/*') if f.endswith('.xls') and pattern in os.path.basename(f)]

        fichiers = [str(f) for f in chemin.glob('*.xls') if (pattern in f.name and number in f.name) ]

        if(len(fichiers) > 1):
            print(f"Multiple files found in dir : {chemin}")
        if(len(fichiers) == 0):
            return []
        return fichiers[0]
    
    
    def fill_buffer(self, chemin) -> int:
        """
        Lit le fichier 1 et remplit le buffer avec les cellules I5 à L13
        """
        # Charger le fichier XLS
        df = pd.read_excel(chemin, usecols='I:', skiprows=4, header=None, engine='xlrd')

        # Sélectionner les données entre les cases I5 à L13
        self.Buffer = pd.DataFrame(data=df.values)
        return EXIT_SUCCESS

    """
    def write_buffer_to_file(self, chemin) -> int:
        # Charger le fichier XLS
        df = pd.read_excel(chemin, header=None)

        #Ecrire le Buffer sur les cases I5 à L13
        df.iloc[4:13, 8:12] = self.Buffer
        df.to_excel(chemin, index=False)

        return EXIT_SUCCESS
    

    
    def write_buffer_to_file(self, chemin) -> int:
        try:            

    

            # Ecrire le Buffer sur les cases I5 à L13
            #df.iloc[4:13, 8:12] = self.Buffer

            # Sauvegarder le DataFrame modifié dans le même fichier
            self.Buffer.to_excel(chemin, index=False, header=False, startrow=4, engine='xlwt')
            #, startcol=8
            return EXIT_SUCCESS

        except Exception as e:
            print(f"Error: {e}")
            return EXIT_FAILED
        """
    

    def write_buffer_to_file(self, chemin) -> int:
        try:
            
            # Open the existing file
            rb = xlrd.open_workbook(chemin, formatting_info=True)

            # Create a copy of the workbook
            wb = copy(rb)

            # Get the first sheet
            sheet = wb.get_sheet(0)

            # Write the Buffer to the specified cells
            for i in range(5, 14):
                for j in range(8, 12):
                    sheet.write(i, j, self.Buffer.iat[i - 5, j - 8])

            # Save the workbook to the specified file
            wb.save(chemin)

            return EXIT_SUCCESS

        except Exception as e:
            print(chemin)
            print(f"Error: {e}")
            return EXIT_FAILED
        

    def convert_xlsx_to_xls(self, xlsx_path, xls_path) -> int:
        try:
            # Load the .xlsx file
            workbook_xlsx = openpyxl.load_workbook(xlsx_path)
            sheet_xlsx = workbook_xlsx.active

            # Create a new .xls file
            workbook_xls = xlwt.Workbook()
            sheet_xls = workbook_xls.add_sheet(sheet_xlsx.title)

            # Copy data from the .xlsx file to the .xls file
            for row_index, row in enumerate(sheet_xlsx.iter_rows()):
                for col_index, cell in enumerate(row):
                    sheet_xls.write(row_index, col_index, cell.value)

            # Save the .xls file
            workbook_xls.save(xls_path)
            
        except:

            return EXIT_FAILED
        return EXIT_SUCCESS
    

    def convert_xls_to_xlsx(self, xls_path, xlsx_path) -> int:
        try:
            # Load the .xls file
            workbook_xls = xlrd.open_workbook(xls_path)
            sheet_xls = workbook_xls.sheet_by_index(0)
            
            # Create a new .xlsx file
            workbook_xlsx = openpyxl.Workbook()
            sheet_xlsx = workbook_xlsx.active
            sheet_xlsx.title = sheet_xls.name
            
            # Copy data from the .xls file to the .xlsx file
            for row_index in range(sheet_xls.nrows):
                for col_index in range(sheet_xls.ncols):
                    cell_value = sheet_xls.cell_value(row_index, col_index)
                    sheet_xlsx.cell(row=row_index + 1, column=col_index + 1, value=cell_value)
            
            # Save the .xlsx file
            workbook_xlsx.save(xlsx_path)
        except:
            return EXIT_FAILED
        return EXIT_SUCCESS
    
        



if(__name__ == "__main__"):

    # variable de vérification
    retour = 0
    def test_retour(retour, message, failed) -> None:
        if(retour == EXIT_SUCCESS):
            print(Fore.GREEN + f'Success : {message}' + Style.RESET_ALL)
        else:
            print(Fore.RED + f"Erreur : {failed}" + Style.RESET_ALL)
    
    document = Document()

    # Création du fichier texte missing file
    retour = document.create_textfile()
    test_retour(retour, "Création du fichier texte", "Création du fichier texte")

    nameXLS = ""

    

    def traitement() -> None:
        #récupérer les infos fichier 1
        retour = document.parse_name(nameXLS)
        if(retour == EXIT_SUCCESS):
            print("\n|===================================|")
            print(Fore.BLUE + f"Doc : {nameXLS}\n" + Style.RESET_ALL)
            document.display_info()

        #chercher fichier 2
        file2Path = document.search_file(document.Strang2Path)

        #chercher fichier 3
        file3Path = document.search_file(document.Strang3Path)

        if(file2Path != [] and file3Path != []):
            #remplir buffer avec fichier 1
            retour = document.fill_buffer(str(document.Strang1Path / Filename.filename))
            test_retour(retour, "filling buffer", "Error : filling buffer")

            #remplir les fichiers 2 et 3 avec le buffer
            retour = document.write_buffer_to_file(file2Path)
            test_retour(retour, "buffer writed to file 2", "Error : writing buffer")

            retour = document.write_buffer_to_file(file3Path)
            test_retour(retour, "buffer writed to file 3", "Error : writing buffer")
            return EXIT_SUCCESS
        
        else:
            print(Fore.RED + "\nErreur : fichiers 2 ou 3 non trouvés" + Style.RESET_ALL)
            print("|__fichier 2 : ", file2Path)
            print("|__fichier 3 : ", file3Path)

            if(file2Path == []):
                document.write_missing_ref(nameXLS, 2)
            if(file3Path == []):
                document.write_missing_ref(nameXLS, 3)

            return EXIT_FAILED



    for file in os.listdir(document.Strang1Path):
            if file.endswith(".xls"):
                #nameXLS = "filename_01_11BEOX12.xls"
                nameXLS = os.path.join(document.Strang1Path, file)
                traitement()


            

        




        
        

